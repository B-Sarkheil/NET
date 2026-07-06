import os
import sys
import glob
import threading
import webbrowser
import openpyxl
from flask import Flask, render_template, jsonify, request, send_file

def get_base_dir():
    """Folder containing the exe (or script)."""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def get_resource_dir():
    """Folder containing templates and static — inside the bundle or next to the script."""
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR     = get_base_dir()
RESOURCE_DIR = get_resource_dir()

app = Flask(__name__,
            template_folder=os.path.join(RESOURCE_DIR, 'templates'),
            static_folder=os.path.join(RESOURCE_DIR,  'static'))

def load_config():
    config = {}
    config_file = os.path.join(BASE_DIR, 'config.txt')
    try:
        with open(config_file, encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if '=' in line:
                    key, _, val = line.partition('=')
                    config[key.strip()] = val.strip()
    except FileNotFoundError:
        pass
    return config

_cfg = load_config()
SOURCE_PATH = _cfg.get('SOURCE_PATH', r"D:\Behnam\NET\fs")
COMMIT_PATH = _cfg.get('COMMIT_PATH', r"D:\Behnam\NET\fs - files")

NEON_COLORS = [
    {'text': '#ff0080', 'border': 'rgba(255,0,128,0.4)',  'glow': 'rgba(255,0,128,0.12)'},
    {'text': '#00ff88', 'border': 'rgba(0,255,136,0.4)',  'glow': 'rgba(0,255,136,0.12)'},
    {'text': '#ff6600', 'border': 'rgba(255,102,0,0.4)',  'glow': 'rgba(255,102,0,0.12)'},
    {'text': '#aa00ff', 'border': 'rgba(170,0,255,0.4)',  'glow': 'rgba(170,0,255,0.12)'},
    {'text': '#ffdc00', 'border': 'rgba(255,220,0,0.4)',  'glow': 'rgba(255,220,0,0.12)'},
    {'text': '#ff3333', 'border': 'rgba(255,51,51,0.4)',  'glow': 'rgba(255,51,51,0.12)'},
]

PROCESS_COLOR = {'text': '#ffb400', 'border': 'rgba(255,180,0,0.45)', 'glow': 'rgba(255,180,0,0.15)'}

def _parse_folder_positions(raw):
    # Format: "Mechanic:right,Instrument:left,Piping:top"
    positions = {}
    for pair in raw.split(','):
        pair = pair.strip()
        if not pair or ':' not in pair:
            continue
        name, _, pos = pair.partition(':')
        name, pos = name.strip(), pos.strip()
        if name and pos:
            positions[name] = pos
    return positions

_DEFAULT_FOLDER_POSITIONS = {
    'Mechanic':  'right',
    'Instrument': 'left',
    'Piping':    'top',
}
FOLDER_POSITIONS = _parse_folder_positions(_cfg['FOLDER_POSITIONS']) if _cfg.get('FOLDER_POSITIONS') else _DEFAULT_FOLDER_POSITIONS

# ── Settings cache (loaded once at startup) ──────────────────────────────────
# Key: "FolderName/SubfolderName"
# Value: dict with all data extracted from setting.xlsx
#   - 'access': list of people (column B of Access sheet)
#   (more fields will be added in future steps)
SETTINGS_CACHE = {}


def load_all_settings():
    # Fallback (Raw Files) Setting.xlsx only feeds Fetch/connections - Access/Notification
    # stay empty so those UI sections stay hidden for folders with no SOURCE_PATH setup.
    cache = {}
    for key in get_all_subfolder_keys():
        folder_name, sub_name = key.split('/', 1)

        primary = glob.glob(os.path.join(SOURCE_PATH, folder_name, sub_name, '[Ss]ource', '[Ss]etting.xlsx'))
        if primary:
            data = parse_setting(primary[0])
            data['has_setting'] = True
            cache[key] = data
            continue

        fallback = glob.glob(os.path.join(COMMIT_PATH, folder_name, sub_name, '[Rr]aw [Ff]iles', '[Ss]etting.xlsx'))
        if fallback:
            data = parse_setting(fallback[0])
            data['access']               = []
            data['notification_headers'] = []
            data['notification_rows']    = []
            data['has_setting']          = False
            cache[key] = data

    return cache


def parse_setting(setting_path):
    data = {
        'access': [], 'ref_paths': [], 'fetch_headers': [], 'fetch_rows_by_ref': {},
        'notification_headers': [], 'notification_rows': [],
        'generate_paths': [], 'generate_headers': [], 'generate_rows_by_ref': {}
    }
    try:
        wb = openpyxl.load_workbook(setting_path, read_only=True, data_only=True)

        if 'Access' in wb.sheetnames:
            ws = wb['Access']
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[1]  # column B
                if val:
                    data['access'].append(str(val))

        if 'Fetch' in wb.sheetnames:
            ws = wb['Fetch']
            rows_iter = ws.iter_rows(values_only=True)

            # Row 1 → headers
            try:
                header_row = next(rows_iter)
                data['fetch_headers'] = [str(h) if h is not None else '' for h in header_row]
            except StopIteration:
                pass

            seen = set()
            for row in rows_iter:
                if len(row) <= 9:
                    continue
                val = row[9]  # column J
                if val:
                    path = str(val).strip()
                    if path not in seen:
                        seen.add(path)
                        data['ref_paths'].append(path)
                    if path not in data['fetch_rows_by_ref']:
                        data['fetch_rows_by_ref'][path] = []
                    data['fetch_rows_by_ref'][path].append(
                        [str(v) if v is not None else '' for v in row]
                    )

        if 'Generator' in wb.sheetnames:
            ws = wb['Generator']
            rows_iter = ws.iter_rows(values_only=True)

            try:
                header_row = next(rows_iter)
                data['generate_headers'] = [str(h) if h is not None else '' for h in header_row]
            except StopIteration:
                pass

            seen = set()
            for row in rows_iter:
                if len(row) <= 6:
                    continue
                val = row[6]  # column G
                if val:
                    path = str(val).strip()
                    if path not in seen:
                        seen.add(path)
                        data['generate_paths'].append(path)
                    if path not in data['generate_rows_by_ref']:
                        data['generate_rows_by_ref'][path] = []
                    data['generate_rows_by_ref'][path].append(
                        [str(v) if v is not None else '' for v in row]
                    )

        if 'Notification' in wb.sheetnames:
            ws = wb['Notification']
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
                data['notification_headers'] = [str(h) if h is not None else '' for h in header_row]
            except StopIteration:
                pass
            for row in rows_iter:
                if any(v is not None for v in row):
                    data['notification_rows'].append(
                        [str(v) if v is not None else '' for v in row]
                    )

        wb.close()
    except Exception:
        pass
    return data


def resolve_path_to_key(ref_path, known_keys):
    # Paths are relative, e.g. "\Process\PDB Approved"
    norm_ref = os.path.normpath(str(ref_path).strip()).lower().lstrip(os.sep)
    parts    = norm_ref.split(os.sep)
    if len(parts) >= 2:
        candidate = f"{parts[0]}/{parts[1]}"
        for key in known_keys:
            if key.lower() == candidate:
                return key
    return None


def get_all_subfolder_keys():
    keys = []
    try:
        for folder_name in sorted(os.listdir(COMMIT_PATH)):
            folder_path = os.path.join(COMMIT_PATH, folder_name)
            if not os.path.isdir(folder_path):
                continue
            for sub_name in sorted(os.listdir(folder_path)):
                if os.path.isdir(os.path.join(folder_path, sub_name)):
                    keys.append(f"{folder_name}/{sub_name}")
    except Exception:
        pass
    return keys


def build_connections(cache):
    connections = []
    all_keys = get_all_subfolder_keys()

    # Fetch sheet: this folder pulls data FROM the referenced folder
    for to_key, data in cache.items():
        for ref_path in data.get('ref_paths', []):
            from_key = resolve_path_to_key(ref_path, all_keys)
            if from_key and from_key != to_key:
                conn = {'from': from_key, 'to': to_key, 'type': 'Fetch'}
                if conn not in connections:
                    connections.append(conn)

    # Generator sheet: this folder produces output TO the referenced folder
    for from_key, data in cache.items():
        for gen_path in data.get('generate_paths', []):
            to_key = resolve_path_to_key(gen_path, all_keys)
            if to_key and to_key != from_key:
                conn = {'from': from_key, 'to': to_key, 'type': 'Generate'}
                if conn not in connections:
                    connections.append(conn)

    return connections


# ── Folder structure ──────────────────────────────────────────────────────────

def get_structure():
    structure = []
    try:
        entries = sorted(os.listdir(COMMIT_PATH))
    except (PermissionError, FileNotFoundError, NotADirectoryError):
        return structure

    for name in entries:
        full_path = os.path.join(COMMIT_PATH, name)
        if os.path.isdir(full_path):
            subfolders = []
            try:
                for sub in sorted(os.listdir(full_path)):
                    sub_path = os.path.join(full_path, sub)
                    if os.path.isdir(sub_path):
                        subfolders.append({'name': sub})
            except (PermissionError, FileNotFoundError, NotADirectoryError):
                pass
            structure.append({'name': name, 'subfolders': subfolders})

    return structure


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    structure = get_structure()

    process = next((f for f in structure if f['name'].lower() == 'process'), None)
    others  = [f for f in structure if f['name'].lower() != 'process']

    for i, folder in enumerate(others):
        folder['color'] = NEON_COLORS[i % len(NEON_COLORS)]

    if process:
        process['color'] = PROCESS_COLOR

    zones = {'top': [], 'bottom': [], 'left': [], 'right': []}
    for folder in others:
        pos = FOLDER_POSITIONS.get(folder['name'])
        if pos in zones:
            zones[pos].append(folder)

    has_middle = bool(zones['left']) or bool(zones['right']) or bool(process)
    row_count = sum([bool(zones['top']), has_middle, bool(zones['bottom'])]) or 1

    return render_template('index.html',
                           process=process,
                           top=zones['top'],
                           right=zones['right'],
                           bottom=zones['bottom'],
                           left=zones['left'],
                           row_count=row_count,
                           has_middle=has_middle,
                           root=SOURCE_PATH)


def _find_nested_setting(parent, sub, subpath):
    """Look up a Setting.xlsx belonging to a folder nested inside a COMMIT_PATH
    subfolder (e.g. Pump DS/P-2001/Source/Setting.xlsx) - used only for that
    nested folder's own Access/Notification, never for connections."""
    folder_path = _safe_subpath(os.path.join(COMMIT_PATH, parent, sub), subpath)
    matches = glob.glob(os.path.join(folder_path, '[Ss]ource', '[Ss]etting.xlsx'))
    if not matches:
        return {}, False
    return parse_setting(matches[0]), True


@app.route('/access')
def get_access():
    parent  = request.args.get('parent', '')
    sub     = request.args.get('sub', '')
    subpath = request.args.get('path', '')

    if not parent or not sub:
        return jsonify({'people': [], 'error': 'Missing parameters'})

    if subpath:
        data, has_setting = _find_nested_setting(parent, sub, subpath)
    else:
        data = SETTINGS_CACHE.get(f"{parent}/{sub}", {})
        has_setting = data.get('has_setting', False)

    return jsonify({'people': data.get('access', []), 'has_setting': has_setting})


@app.route('/debug-fetch')
def debug_fetch():
    parent = request.args.get('parent', '')
    sub    = request.args.get('sub', '')
    matches = glob.glob(os.path.join(SOURCE_PATH, parent, sub, '[Ss]ource', '[Ss]etting.xlsx'))
    if not matches:
        return jsonify({'error': 'File not found'})
    wb = openpyxl.load_workbook(matches[0], read_only=True, data_only=True)
    if 'Fetch' not in wb.sheetnames:
        return jsonify({'error': 'No Fetch sheet'})
    ws = wb['Fetch']
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        rows.append({'row': i, 'col_J': row[9] if len(row) > 9 else 'N/A'})
    wb.close()
    return jsonify({'rows': rows})


@app.route('/reload')
def reload_cache():
    global SETTINGS_CACHE, CONNECTIONS
    SETTINGS_CACHE = load_all_settings()
    CONNECTIONS    = build_connections(SETTINGS_CACHE)
    return jsonify({
        'status': 'ok',
        'cache_keys': list(SETTINGS_CACHE.keys()),
        'connections': CONNECTIONS,
    })


@app.route('/debug')
def debug():
    result = {}
    for key, data in SETTINGS_CACHE.items():
        result[key] = {
            'access': data.get('access', []),
            'ref_paths': data.get('ref_paths', []),
        }
    return jsonify({
        'source_path': SOURCE_PATH,
        'cache_keys': list(SETTINGS_CACHE.keys()),
        'connections': CONNECTIONS,
        'details': result,
    })


@app.route('/connections')
def get_connections():
    return jsonify({'connections': CONNECTIONS})


def _safe_subpath(base_path, subpath):
    """Join base_path with a user-supplied relative subpath, rejecting any
    segment that could escape base_path (.., absolute paths, empty)."""
    parts = [p for p in subpath.replace('\\', '/').split('/') if p not in ('', '.', '..')]
    return os.path.join(base_path, *parts) if parts else base_path


@app.route('/files')
def get_files():
    parent  = request.args.get('parent', '')
    sub     = request.args.get('sub', '')
    subpath = request.args.get('path', '')

    if not parent or not sub:
        return jsonify({'files': [], 'error': 'Missing parameters'})

    folder_path = _safe_subpath(os.path.join(COMMIT_PATH, parent, sub), subpath)

    try:
        entries = sorted(os.listdir(folder_path))
        files = [
            {'name': e, 'is_dir': os.path.isdir(os.path.join(folder_path, e))}
            for e in entries
            if not (os.path.isdir(os.path.join(folder_path, e)) and e.lower() == 'raw files')
        ]
        return jsonify({'files': files, 'folder': folder_path})
    except PermissionError:
        return jsonify({'files': [], 'error': 'Permission denied'})
    except FileNotFoundError:
        return jsonify({'files': [], 'error': 'Folder not found'})


# ── Startup ───────────────────────────────────────────────────────────────────

SETTINGS_CACHE = load_all_settings()
CONNECTIONS    = build_connections(SETTINGS_CACHE)

@app.route('/notification')
def get_notification():
    parent  = request.args.get('parent', '')
    sub     = request.args.get('sub', '')
    subpath = request.args.get('path', '')
    if not parent or not sub:
        return jsonify({'headers': [], 'rows': [], 'error': 'Missing parameters'})

    if subpath:
        data, has_setting = _find_nested_setting(parent, sub, subpath)
    else:
        data = SETTINGS_CACHE.get(f"{parent}/{sub}", {})
        has_setting = data.get('has_setting', False)

    return jsonify({
        'headers': data.get('notification_headers', []),
        'rows':    data.get('notification_rows', []),
        'has_setting': has_setting,
    })


@app.route('/fetch-data')
def get_fetch_data():
    from_key    = request.args.get('from_key', '')
    to_parent   = request.args.get('to_parent', '')
    to_sub      = request.args.get('to_sub', '')

    to_key = f"{to_parent}/{to_sub}"
    data   = SETTINGS_CACHE.get(to_key, {})

    all_keys      = get_all_subfolder_keys()
    matching_rows = []

    for ref_path, rows in data.get('fetch_rows_by_ref', {}).items():
        if resolve_path_to_key(ref_path, all_keys) == from_key:
            matching_rows.extend(rows)

    return jsonify({
        'headers': data.get('fetch_headers', []),
        'rows':    matching_rows,
        'from':    from_key,
        'to':      to_key,
    })


@app.route('/generate-data')
def get_generate_data():
    from_parent = request.args.get('from_parent', '')
    from_sub    = request.args.get('from_sub', '')
    to_key      = request.args.get('to_key', '')

    from_key = f"{from_parent}/{from_sub}"
    data     = SETTINGS_CACHE.get(from_key, {})

    all_keys      = get_all_subfolder_keys()
    matching_rows = []

    for gen_path, rows in data.get('generate_rows_by_ref', {}).items():
        if resolve_path_to_key(gen_path, all_keys) == to_key:
            matching_rows.extend(rows)

    return jsonify({
        'headers': data.get('generate_headers', []),
        'rows':    matching_rows,
        'from':    from_key,
        'to':      to_key,
    })


@app.route('/download')
def download_file():
    parent   = request.args.get('parent', '')
    sub      = request.args.get('sub', '')
    subpath  = request.args.get('path', '')
    filename = request.args.get('file', '')

    if not parent or not sub or not filename or os.path.basename(filename) != filename:
        return 'Missing parameters', 400

    folder_path = _safe_subpath(os.path.join(COMMIT_PATH, parent, sub), subpath)
    file_path   = os.path.join(folder_path, filename)

    if not os.path.isfile(file_path):
        return 'File not found', 404

    return send_file(file_path, as_attachment=True, download_name=filename)


if __name__ == '__main__':
    threading.Timer(1.5, lambda: webbrowser.open('http://localhost:5000')).start()
    app.run(host='0.0.0.0', port=5000, debug=True, use_reloader=False)
